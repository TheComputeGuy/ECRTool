import openpyxl as oxl
from tkinter import filedialog
import tkinter

filename=''
dirname=''

root = tkinter.Tk()
root.title("ECRTool")

tkinter.Label(root, text="Choose XLSX file: ").grid(row=0)
def filebrowsefunc():
    global filename
    filename = filedialog.askopenfilename(filetypes=[("XLSX files", "*.xlsx")])
    filepathlabel.config(text=filename)
filebrowsebutton = tkinter.Button(root, text="Browse", command=filebrowsefunc)
filebrowsebutton.grid(row=0, column=1)
filepathlabel = tkinter.Label(root)
filepathlabel.grid(row=0, column=2)

tkinter.Label(root, text="Select save folder: ").grid(row=1)
def dirbrowsefunc():
    global dirname
    dirname = filedialog.askdirectory()
    dirpathlabel.config(text=dirname)
dirbrowsebutton = tkinter.Button(root, text="Browse", command=dirbrowsefunc)
dirbrowsebutton.grid(row=1, column=1)
dirpathlabel = tkinter.Label(root)
dirpathlabel.grid(row=1, column=2)

tkinter.Label(root, text="Enter month and year of ECR: ").grid(row=2)
month_entry = tkinter.Entry(root)
month_entry.grid(row=2, column=1)


def conv():
    global filename
    global dirname

    xlsxname = filename
    month = month_entry.get()
    time = month.strip().replace(" ", "_")
    txtfilename = dirname+"\\"+time+".txt"          #Changes from *nix version to Windows
    txtfile = open(txtfilename, "a")

    wb = oxl.load_workbook(xlsxname, data_only=True)
    sheet = wb.active

    for rowno in range(2, sheet.max_row+1):
        val = sheet.cell(row=rowno, column=3).value
        if(val != 0):
            vals = []
            for k in range(1, 12):
                val = sheet.cell(row=rowno, column=k).value
                if(isinstance(val, float)):
                    val = int(val)
                if(isinstance(val, int) and val == 0):
                    val = ''
                vals.append(str(val))
            txtfile.write("#~#".join(vals))
            txtfile.write('\n')
    txtfile.close()
    statuslabel.config(text="File converted and saved successfully!", bg="green")
    

submitbutton = tkinter.Button(root, text="Submit and convert", command=conv)
submitbutton.grid(row=3, column=1)
statuslabel = tkinter.Label(root)
statuslabel.grid(row=4, column=1)


if __name__ == "__main__":
    root.mainloop()
